#!/usr/bin/env python3
"""
Build the Project Strategy TOR (Terms of Reference) Dashboard workbook.

Team reporting pack that tracks live project delivery against what each
project's Terms of Reference actually committed to: scope, deliverables,
milestones, budget, governance (RACI) and risk.

Two cadences share one dashboard:
  Daily   - fine-grained tracking during the month, driven by 'Daily Log'
  Monthly - the governance / SteerCo view, driven by 'Monthly Log'
The Dashboard's Cadence cell switches between them; every metric follows.

Run:  python3 build_tor_dashboard.py
Then: python3 <xlsx-skill>/scripts/recalc.py Project_Strategy_TOR_Dashboard.xlsx

Regenerating overwrites the workbook, so edit this script rather than the .xlsx
if you want to change structure; edit the .xlsx directly for day-to-day data.
"""

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "Project_Strategy_TOR_Dashboard.xlsx")

# ---------------------------------------------------------------- styling ----
FONT = "Arial"

NAVY = "1F3864"
SLATE = "44546A"
LIGHT = "D9E2F3"
BAND = "F2F5FB"
GREY = "808080"

BLUE_INPUT = "0000FF"   # hardcoded input typed by the user
GREEN_LINK = "008000"   # value pulled from another sheet
BLACK = "000000"        # formula computed on this sheet
YELLOW_FILL = "FFFF00"  # cells the user is expected to fill in

RAG_FILL = {
    "Red": PatternFill("solid", fgColor="F4B7B7"),
    "Amber": PatternFill("solid", fgColor="FFE39A"),
    "Green": PatternFill("solid", fgColor="C2E0C2"),
}
SEVERITY_FILL = {
    "High": RAG_FILL["Red"],
    "Medium": RAG_FILL["Amber"],
    "Low": RAG_FILL["Green"],
}

CUR = '$#,##0;($#,##0);-'
PCT = '0.0%;(0.0%);-'
INT = '#,##0;(#,##0);-'
DATE = 'yyyy-mm-dd'
MONTH = 'yyyy-mm'

thin = Side(style="thin", color="BFBFBF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

# Sheet-row capacity. Formulas span these ranges so added rows are picked up
# automatically without anyone editing a formula.
TOR_FIRST, TOR_LAST = 4, 43        # 40 projects
LOG_FIRST, LOG_LAST = 4, 303       # 300 monthly rows
DAY_FIRST, DAY_LAST = 4, 1503      # 1500 daily rows
DEL_FIRST, DEL_LAST = 4, 203
RSK_FIRST, RSK_LAST = 4, 203


def title_block(ws, title, subtitle, width):
    ws["A1"] = title
    ws["A1"].font = Font(name=FONT, size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(vertical="center", indent=1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    ws.row_dimensions[1].height = 26

    ws["A2"] = subtitle
    ws["A2"].font = Font(name=FONT, size=9, italic=True, color=SLATE)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)


def header_row(ws, row, headers, wrap=True):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=SLATE)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=wrap)
        c.border = BOX
    ws.row_dimensions[row].height = 32


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


def body(cell, *, fmt=None, color=BLACK, bold=False, wrap=False,
         align=None, fill=None):
    cell.font = Font(name=FONT, size=9, color=color, bold=bold)
    cell.border = BOX
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal=align, vertical="top" if wrap else "center",
                               wrap_text=wrap)


def band(ws, first, last, ncols):
    """Zebra-stripe alternate rows so long registers stay readable."""
    for r in range(first, last + 1):
        if (r - first) % 2 == 1:
            for c in range(1, ncols + 1):
                cell = ws.cell(row=r, column=c)
                if cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = PatternFill("solid", fgColor=BAND)


# ------------------------------------------------------------ example data ---
# EXAMPLE PORTFOLIO - replace with the team's real projects. Names are
# placeholders, not real people.
MONTHS = [dt.date(2026, 3, 1), dt.date(2026, 4, 1), dt.date(2026, 5, 1),
          dt.date(2026, 6, 1), dt.date(2026, 7, 1), dt.date(2026, 8, 1)]
REPORT_MONTH = MONTHS[-1]

# Daily example data covers the working days of the final month, so the daily
# and monthly views can be compared directly.
DAILY_DAYS = [d for d in (dt.date(2026, 8, 1) + dt.timedelta(days=i) for i in range(31))
              if d.weekday() < 5]
REPORT_DAY = DAILY_DAYS[-1]

PROJECTS = [
    # id, name, pillar, sponsor, pm, tor approved, start, end, budget,
    # planned milestones, planned deliverables, objective, in scope, out of scope
    ("P-001", "Client Portal Rebuild", "Digital Experience", "A. Mensah",
     "R. Patel", dt.date(2026, 2, 10), dt.date(2026, 3, 1), dt.date(2026, 11, 30),
     500000, 9, 12,
     "Replace the legacy client portal to cut self-service failure rate by half.",
     "Portal UI, auth, self-service billing, migration of active accounts",
     "Mobile app, partner API, archived account history"),
    ("P-002", "Data Platform Migration", "Data & Infrastructure", "L. Njoku",
     "S. Kowalski", dt.date(2026, 2, 3), dt.date(2026, 3, 1), dt.date(2027, 1, 29),
     820000, 12, 15,
     "Move reporting workloads off the on-prem warehouse to the cloud platform.",
     "ETL rebuild, warehouse cutover, BI re-point, decommission plan",
     "ML feature store, real-time streaming, third-party data contracts"),
    ("P-003", "Regulatory Reporting Uplift", "Risk & Compliance", "D. Ferreira",
     "M. Haddad", dt.date(2026, 2, 17), dt.date(2026, 3, 1), dt.date(2026, 10, 30),
     340000, 8, 10,
     "Meet the new quarterly disclosure standard ahead of the Q1 2027 deadline.",
     "Control redesign, evidence automation, submission templates, sign-off flow",
     "Historic restatements, group consolidation changes"),
    ("P-004", "Market Expansion - APAC", "Growth", "A. Mensah",
     "T. Yamamoto", dt.date(2026, 2, 24), dt.date(2026, 3, 1), dt.date(2027, 2, 26),
     610000, 10, 11,
     "Stand up sales and support operations in two APAC markets.",
     "Entity setup, local hiring, pricing localisation, support coverage",
     "Manufacturing footprint, local data centre, M&A options"),
    ("P-005", "Cost-to-Serve Optimisation", "Operating Efficiency", "L. Njoku",
     "C. Duarte", dt.date(2026, 2, 6), dt.date(2026, 3, 1), dt.date(2026, 9, 30),
     275000, 7, 9,
     "Reduce cost to serve per account by 15% without lowering CSAT.",
     "Process redesign, tier-1 automation, vendor renegotiation",
     "Headcount reduction, offshoring, product rationalisation"),
]

# Cumulative-to-date figures per project, one value per month in MONTHS.
LOG = {
    "P-001": dict(
        ms_plan=[2, 3, 5, 6, 8, 9],       ms_done=[2, 3, 4, 5, 7, 7],
        dl_due=[2, 4, 5, 7, 9, 11],       dl_acc=[2, 4, 4, 6, 8, 9],
        spend=[72000, 140000, 205000, 286000, 352000, 412000],
        cr_raised=[0, 1, 1, 2, 3, 3],     cr_appr=[0, 1, 1, 1, 2, 2],
        risks=[3, 4, 4, 5, 4, 3],         high=[0, 1, 1, 1, 1, 1],
        gaps=[1, 1, 0, 0, 2, 2],
        rag=["Green", "Green", "Amber", "Amber", "Amber", "Amber"],
        note=["Kick-off complete, TOR baselined.",
              "Auth vendor confirmed; first CR raised on billing scope.",
              "Design sign-off slipped 2 weeks; one milestone missed.",
              "Migration dry-run failed once, retest scheduled.",
              "Two deliverables awaiting sponsor acceptance.",
              "Recovery plan agreed; hit rate improving but still behind TOR."],
        daily_note="Recovery plan tasks progressing; auth cutover in test."),
    "P-002": dict(
        ms_plan=[1, 2, 4, 6, 8, 10],      ms_done=[1, 2, 4, 6, 7, 9],
        dl_due=[1, 3, 4, 6, 8, 10],       dl_acc=[1, 3, 4, 5, 7, 9],
        spend=[95000, 190000, 300000, 430000, 560000, 690000],
        cr_raised=[1, 1, 2, 2, 2, 3],     cr_appr=[0, 1, 1, 1, 1, 2],
        risks=[5, 5, 6, 6, 5, 5],         high=[1, 1, 2, 1, 1, 0],
        gaps=[2, 1, 1, 0, 0, 0],
        rag=["Amber", "Amber", "Amber", "Green", "Green", "Green"],
        note=["Two RACI gaps flagged at TOR review.",
              "Ownership resolved for ETL workstream.",
              "Cutover risk raised to high pending capacity test.",
              "Capacity test passed; back on plan.",
              "One milestone slipped into next month, no scope impact.",
              "Tracking to TOR; decommission CR approved."],
        daily_note="Cutover runbook review and BI parity testing under way."),
    "P-003": dict(
        ms_plan=[2, 3, 4, 5, 7, 8],       ms_done=[2, 3, 3, 3, 4, 5],
        dl_due=[2, 3, 4, 6, 8, 10],       dl_acc=[2, 3, 3, 4, 5, 6],
        spend=[58000, 112000, 168000, 232000, 296000, 351000],
        cr_raised=[0, 0, 1, 3, 4, 5],     cr_appr=[0, 0, 1, 2, 3, 4],
        risks=[4, 5, 7, 8, 9, 9],         high=[1, 1, 2, 3, 3, 3],
        gaps=[0, 1, 2, 3, 3, 4],
        rag=["Green", "Amber", "Amber", "Red", "Red", "Red"],
        note=["Baselined against draft standard.",
              "Regulator clarification pending, first slip.",
              "Scope grew: evidence automation added by CR.",
              "Budget forecast breached; escalated to sponsor.",
              "Four unowned deliverables, accountability unresolved.",
              "Overspent vs TOR budget. Re-baseline decision needed at SteerCo."],
        daily_note="Scope freeze in force pending the re-baseline paper."),
    "P-004": dict(
        ms_plan=[1, 2, 3, 5, 6, 8],       ms_done=[1, 2, 3, 5, 6, 8],
        dl_due=[1, 2, 3, 4, 6, 8],        dl_acc=[1, 2, 3, 4, 6, 8],
        spend=[60000, 125000, 190000, 268000, 340000, 415000],
        cr_raised=[0, 0, 0, 1, 1, 1],     cr_appr=[0, 0, 0, 1, 1, 1],
        risks=[2, 2, 3, 3, 2, 2],         high=[0, 0, 0, 0, 0, 0],
        gaps=[0, 0, 0, 0, 0, 0],
        rag=["Green", "Green", "Green", "Green", "Green", "Green"],
        note=["Entity setup started on schedule.",
              "Local counsel engaged, no issues.",
              "Hiring pipeline ahead of plan.",
              "Pricing CR approved, budget unchanged.",
              "First market live.",
              "Both markets on track to TOR end date."],
        daily_note="Second market onboarding on plan."),
    "P-005": dict(
        ms_plan=[1, 2, 3, 4, 6, 7],       ms_done=[1, 2, 3, 4, 5, 6],
        dl_due=[1, 2, 3, 5, 7, 9],        dl_acc=[1, 2, 3, 4, 6, 7],
        spend=[30000, 62000, 95000, 138000, 180000, 224000],
        cr_raised=[0, 1, 1, 1, 2, 2],     cr_appr=[0, 0, 0, 0, 1, 1],
        risks=[3, 3, 3, 4, 4, 3],         high=[0, 0, 1, 1, 1, 0],
        gaps=[1, 1, 1, 0, 0, 1],
        rag=["Green", "Green", "Amber", "Amber", "Green", "Green"],
        note=["Baseline agreed with operations.",
              "Vendor renegotiation opened.",
              "Automation vendor risk raised to high.",
              "Mitigation in place, second vendor shortlisted.",
              "Vendor risk closed; CR approved for extra automation.",
              "Two deliverables pending acceptance, otherwise on plan."],
        daily_note="Tier-1 automation release in acceptance testing."),
}


def ramp(start, end, n, monotonic=True):
    """Spread a month's movement across n days; the last value lands on `end`.

    Cumulative measures only ever rise, so they are forced monotonic. Snapshot
    measures (risk and RACI counts) can fall, so they are left to move freely.
    """
    vals = []
    for i in range(n):
        t = (i + 1) / n
        vals.append(round(start + (end - start) * t))
    vals[-1] = end
    if monotonic:
        for i in range(1, n):
            vals[i] = max(vals[i], vals[i - 1])
    return vals


def build_daily(pid):
    """Daily rows for the final month, converging on that month's monthly row."""
    d = LOG[pid]
    n = len(DAILY_DAYS)
    prev, last = -2, -1          # previous month's figure -> final month's figure
    series = {}
    for key, mono in (("ms_plan", True), ("ms_done", True), ("dl_due", True),
                      ("dl_acc", True), ("spend", True), ("cr_raised", True),
                      ("cr_appr", True), ("risks", False), ("high", False),
                      ("gaps", False)):
        series[key] = ramp(d[key][prev], d[key][last], n, monotonic=mono)
    series["rag"] = [d["rag"][last]] * n
    series["note"] = [d["daily_note"]] * (n - 1) + [d["note"][last]]
    return series


DAILY = {p[0]: build_daily(p[0]) for p in PROJECTS}

DELIVERABLES = [
    # id, project, deliverable, TOR ref, R, A, C, I, due, status
    ("D-001", "P-001", "Portal design system sign-off", "TOR 4.1", "R. Patel", "A. Mensah", "Design guild", "Steering group", dt.date(2026, 4, 30), "Accepted"),
    ("D-002", "P-001", "Authentication service cutover", "TOR 4.2", "K. Osei", "R. Patel", "Security", "Service desk", dt.date(2026, 7, 31), "Submitted"),
    ("D-003", "P-001", "Self-service billing module", "TOR 4.3", "K. Osei", "", "Finance ops", "Client services", dt.date(2026, 9, 30), "In Progress"),
    ("D-004", "P-001", "Active account migration", "TOR 4.4", "J. Silva", "", "Data team", "Steering group", dt.date(2026, 10, 30), "Not Started"),
    ("D-005", "P-002", "ETL rebuild - phase 1", "TOR 3.1", "S. Kowalski", "L. Njoku", "Platform", "BI users", dt.date(2026, 5, 29), "Accepted"),
    ("D-006", "P-002", "Warehouse cutover runbook", "TOR 3.2", "N. Abebe", "S. Kowalski", "Ops", "All staff", dt.date(2026, 8, 31), "Submitted"),
    ("D-007", "P-002", "BI re-point and validation", "TOR 3.3", "N. Abebe", "S. Kowalski", "Finance", "BI users", dt.date(2026, 10, 30), "In Progress"),
    ("D-008", "P-003", "Control redesign pack", "TOR 2.1", "M. Haddad", "D. Ferreira", "Internal audit", "Board", dt.date(2026, 5, 29), "Accepted"),
    ("D-009", "P-003", "Evidence automation build", "TOR 2.4 (CR-03)", "P. Lindqvist", "", "Compliance", "Board", dt.date(2026, 8, 31), "In Progress"),
    ("D-010", "P-003", "Submission template set", "TOR 2.2", "P. Lindqvist", "", "Regulator liaison", "Board", dt.date(2026, 9, 30), "Not Started"),
    ("D-011", "P-003", "Sign-off workflow", "TOR 2.3", "M. Haddad", "", "Legal", "Board", dt.date(2026, 10, 30), "Not Started"),
    ("D-012", "P-003", "Historic restatement review", "TOR 2.5 (CR-04)", "M. Haddad", "", "Finance", "Board", dt.date(2026, 10, 30), "Not Started"),
    ("D-013", "P-004", "APAC entity registration", "TOR 5.1", "T. Yamamoto", "A. Mensah", "Legal", "Exec", dt.date(2026, 5, 29), "Accepted"),
    ("D-014", "P-004", "Localised pricing model", "TOR 5.2", "H. Tran", "T. Yamamoto", "Finance", "Sales", dt.date(2026, 7, 31), "Accepted"),
    ("D-015", "P-004", "Regional support coverage", "TOR 5.3", "H. Tran", "T. Yamamoto", "Support", "Clients", dt.date(2026, 11, 30), "In Progress"),
    ("D-016", "P-005", "Process redesign blueprint", "TOR 6.1", "C. Duarte", "L. Njoku", "Ops leads", "Exec", dt.date(2026, 5, 29), "Accepted"),
    ("D-017", "P-005", "Tier-1 automation release", "TOR 6.2", "E. Boateng", "C. Duarte", "Service desk", "Ops", dt.date(2026, 8, 31), "Submitted"),
    ("D-018", "P-005", "Vendor contract renegotiation", "TOR 6.3", "E. Boateng", "", "Procurement", "Finance", dt.date(2026, 9, 30), "In Progress"),
]

RISKS = [
    # id, project, type, description, likelihood, impact, owner, mitigation, status, raised, target
    ("R-001", "P-001", "Risk", "Auth vendor cannot meet the agreed cutover window", 3, 4, "R. Patel", "Parallel-run window booked; fallback to extended legacy support", "Mitigating", dt.date(2026, 4, 14), dt.date(2026, 9, 30)),
    ("R-002", "P-001", "Issue", "Design sign-off slipped two weeks against the TOR schedule", 5, 3, "R. Patel", "Recovery plan agreed with sponsor; two milestones re-sequenced", "Open", dt.date(2026, 5, 12), dt.date(2026, 9, 15)),
    ("R-003", "P-001", "Risk", "Billing scope creep beyond TOR 4.3 without an approved CR", 3, 3, "A. Mensah", "CR-02 tabled at SteerCo for decision", "Open", dt.date(2026, 7, 8), dt.date(2026, 9, 30)),
    ("R-004", "P-002", "Risk", "Cutover exceeds the agreed maintenance window", 2, 5, "S. Kowalski", "Capacity test passed; rehearsal booked for October", "Mitigating", dt.date(2026, 5, 6), dt.date(2026, 10, 15)),
    ("R-005", "P-002", "Risk", "Legacy warehouse licence renews before decommission", 3, 3, "L. Njoku", "Decommission CR approved; renewal deferred to monthly terms", "Closed", dt.date(2026, 4, 2), dt.date(2026, 8, 1)),
    ("R-006", "P-003", "Issue", "Forecast spend exceeds the TOR-approved budget", 5, 5, "D. Ferreira", "Re-baseline paper going to SteerCo; scope freeze in force", "Open", dt.date(2026, 6, 10), dt.date(2026, 9, 15)),
    ("R-007", "P-003", "Issue", "Four deliverables have no accountable owner", 5, 4, "D. Ferreira", "Accountability workshop scheduled with the sponsor", "Open", dt.date(2026, 7, 3), dt.date(2026, 9, 12)),
    ("R-008", "P-003", "Risk", "Regulator clarification may reopen the control design", 3, 4, "M. Haddad", "Formal query submitted; contingency in the plan", "Open", dt.date(2026, 4, 21), dt.date(2026, 10, 30)),
    ("R-009", "P-003", "Risk", "Scope added by CR-03/CR-04 not funded in the TOR", 4, 4, "D. Ferreira", "Funding request bundled with the re-baseline paper", "Open", dt.date(2026, 6, 25), dt.date(2026, 9, 15)),
    ("R-010", "P-004", "Risk", "Local hiring lead times slip in the second market", 2, 3, "T. Yamamoto", "Agency retained; pipeline running two candidates deep", "Mitigating", dt.date(2026, 5, 18), dt.date(2026, 11, 30)),
    ("R-011", "P-004", "Risk", "Currency movement erodes the localised pricing margin", 2, 3, "H. Tran", "Quarterly price review clause added to contracts", "Open", dt.date(2026, 7, 15), dt.date(2026, 12, 18)),
    ("R-012", "P-005", "Risk", "Automation vendor stability after the funding round", 2, 4, "C. Duarte", "Second vendor shortlisted; source code escrow agreed", "Closed", dt.date(2026, 5, 20), dt.date(2026, 8, 14)),
    ("R-013", "P-005", "Risk", "Benefit realisation depends on a CSAT floor being held", 3, 3, "C. Duarte", "CSAT tracked weekly; rollback trigger defined", "Mitigating", dt.date(2026, 6, 30), dt.date(2026, 9, 30)),
    ("R-014", "P-002", "Risk", "Source system owners slow to sign off data contracts", 3, 3, "S. Kowalski", "Contract templates pre-agreed; weekly chase in the workstream call", "Open", dt.date(2026, 6, 8), dt.date(2026, 10, 30)),
    ("R-015", "P-002", "Risk", "BI report parity testing may run past the agreed window", 3, 4, "N. Abebe", "Parity suite automated; top 40 reports prioritised", "Mitigating", dt.date(2026, 7, 1), dt.date(2026, 11, 27)),
    ("R-016", "P-002", "Issue", "Two ETL jobs exceed the nightly batch window", 4, 3, "N. Abebe", "Partitioning rework in flight; interim staggered schedule", "Open", dt.date(2026, 7, 22), dt.date(2026, 9, 30)),
    ("R-017", "P-002", "Risk", "Cloud run-rate tracking above the TOR estimate", 2, 4, "S. Kowalski", "Reserved capacity purchased; monthly cost review added", "Mitigating", dt.date(2026, 8, 5), dt.date(2026, 12, 18)),
    ("R-018", "P-003", "Risk", "Evidence automation depends on a single contractor", 3, 4, "M. Haddad", "Knowledge transfer sessions booked; handover pack required", "Mitigating", dt.date(2026, 6, 17), dt.date(2026, 10, 15)),
    ("R-019", "P-003", "Risk", "Template changes require a second legal review", 3, 3, "M. Haddad", "Legal slot pre-booked for the October cycle", "Open", dt.date(2026, 7, 9), dt.date(2026, 10, 16)),
    ("R-020", "P-003", "Issue", "Control testing environment unavailable for two weeks in September", 4, 3, "P. Lindqvist", "Testing re-sequenced; shared environment booked as fallback", "Open", dt.date(2026, 8, 12), dt.date(2026, 9, 26)),
    ("R-021", "P-003", "Risk", "Sign-off workflow tooling licence not yet procured", 3, 3, "D. Ferreira", "Procurement raised; manual workflow as interim", "Open", dt.date(2026, 8, 3), dt.date(2026, 9, 30)),
    ("R-022", "P-003", "Risk", "Reporting calendar clashes with year-end close", 2, 4, "D. Ferreira", "Dates agreed with finance; submission moved one week earlier", "Mitigating", dt.date(2026, 8, 19), dt.date(2026, 11, 30)),
    ("R-023", "P-005", "Risk", "Process redesign adoption varies by region", 3, 3, "C. Duarte", "Regional champions appointed; adoption tracked weekly", "Open", dt.date(2026, 7, 14), dt.date(2026, 9, 30)),
    ("R-024", "P-005", "Risk", "Benefit baseline disputed by finance", 2, 4, "C. Duarte", "Baseline restated with finance and re-signed by the sponsor", "Mitigating", dt.date(2026, 8, 7), dt.date(2026, 9, 26)),
]

LIST_RAG = ["Green", "Amber", "Red"]
LIST_DSTATUS = ["Not Started", "In Progress", "Submitted", "Accepted", "Descoped"]
LIST_RSTATUS = ["Open", "Mitigating", "Closed"]
LIST_RTYPE = ["Risk", "Issue"]
LIST_SCORE = [1, 2, 3, 4, 5]
LIST_CADENCE = ["Daily", "Monthly"]

wb = Workbook()

# =============================================================== READ ME =====
ws = wb.active
ws.title = "Read Me"
ws.sheet_properties.tabColor = NAVY
ws.sheet_view.showGridLines = False
widths(ws, {"A": 3, "B": 26, "C": 96})

title_block(ws, "  Project Strategy TOR Dashboard - daily and monthly",
            "  Tracks delivery against what each project's Terms of Reference committed to. "
            "Log daily during the month; report monthly at governance.", 4)

r = 4


def section(text):
    global r
    c = ws.cell(row=r, column=2, value=text)
    c.font = Font(name=FONT, size=11, bold=True, color=NAVY)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws.cell(row=r, column=2).border = Border(bottom=Side(style="medium", color=LIGHT))
    r += 1


def line(label, text, label_color=SLATE):
    global r
    a = ws.cell(row=r, column=2, value=label)
    a.font = Font(name=FONT, size=9, bold=True, color=label_color)
    a.alignment = Alignment(vertical="top", wrap_text=True)
    b = ws.cell(row=r, column=3, value=text)
    b.font = Font(name=FONT, size=9)
    b.alignment = Alignment(vertical="top", wrap_text=True)
    r += 1


def gap(n=1):
    global r
    r += n


section("What this workbook is for")
line("", "A single pack the team can run at two speeds. Every metric answers one question: "
         "is this project still delivering what its Terms of Reference said it would, for the "
         "money and by the dates the TOR approved?")
gap()

section("Two cadences, one dashboard")
line("Daily", "Log progress each working day on 'Daily Log'. Use it to spot slippage inside the "
              "month, while there is still time to act. Set Dashboard C3 to Daily and put the "
              "date in C4.")
line("Monthly", "One row per project per month on 'Monthly Log' - the governance and SteerCo "
                "record, and what the monthly trend charts keep. Set Dashboard C3 to Monthly and "
                "put the first day of the month in C4.")
line("How they join", "Both logs hold the same cumulative-to-date figures, so the last daily row "
                      "of a month IS that month's monthly row. The roll-up block on 'Daily Log' "
                      "reads off any date so you can copy it straight into the Monthly Log - no "
                      "double entry, no drift.")
gap()

section("The daily routine - two minutes per project")
line("1.  Add today's rows", "On 'Daily Log', copy yesterday's block down, change the date, and "
                             "update anything that moved. Figures are CUMULATIVE TO DATE, not "
                             "in-day.")
line("2.  Look at the Dashboard", "Cadence Daily, today's date in C4. Anything Red or Amber is "
                                  "today's problem, not next month's.")
gap()

section("The monthly cycle - five steps")
line("1.  Update registers",
     "Add or update rows on 'Deliverables & RACI' and 'Risks & Issues' as things change.")
line("2.  Roll the month up",
     "On 'Daily Log', put the month-end date in the roll-up block (cell R3). It shows each "
     "project's figures on that date. Copy them into a new month row on 'Monthly Log'. If the "
     "team does not log daily, just type the month's figures straight into 'Monthly Log'.")
line("3.  Copy the live counts",
     "Copy Open Risks and High Risks from 'Risks & Issues', and RACI Gaps from "
     "'Deliverables & RACI', into the month's row. Freezing them into the log is what gives the "
     "trend charts their history - the registers only ever show today.")
line("4.  Set cadence and month",
     "On 'Dashboard', set C3 to Monthly and C4 to the first day of the month.")
line("5.  Review and present",
     "Walk the Dashboard exception table, then 'Charts' for the trend. Red and Amber rows are "
     "the agenda.")
gap()

section("Tab guide")
line("Read Me", "This page.")
line("CEO Brief", "The five things worth knowing, in sentences. Every line is calculated from "
                  "the Dashboard at the cadence and date set there, so it cannot be edited into "
                  "a better story. This is what the file opens on.")
line("TOR Register", "The baseline. One row per project holding what the signed TOR committed to: "
                     "objective, scope boundaries, dates, approved budget, planned milestone and "
                     "deliverable counts. Change this only through an approved change request.")
line("Daily Log", "Day-by-day entry, one row per project per working day, cumulative to date. "
                  "Also carries the month-end roll-up block.")
line("Monthly Log", "One row per project per month, cumulative to date. The governance record "
                    "and the source of the monthly trend charts.")
line("Deliverables & RACI", "Deliverable-level register mapped back to TOR clauses, with "
                            "Responsible / Accountable / Consulted / Informed. A blank Accountable "
                            "is a governance gap and is counted for you.")
line("Risks & Issues", "Risk and issue register with likelihood x impact scoring and live counts "
                       "per project.")
line("Dashboard", "Portfolio tiles and a per-project table, for whichever cadence C3 is set to.")
line("Charts", "Monthly trend, per-project comparisons, and daily trend for the current month.")
line("Lists", "Dropdown values. Extend here if the team uses different status wording.")
gap()

section("How each metric is defined")
line("Scope adherence %", "Deliverables accepted to date / deliverables due to date. Measures "
                          "whether the TOR's promised outputs are actually landing and being "
                          "accepted, not just started.")
line("Milestone hit rate %", "Milestones achieved to date / milestones planned to date. Schedule "
                             "health against the TOR baseline.")
line("Budget used %", "Spend to date / TOR-approved budget. Over 100% means the project has spent "
                      "past what the TOR authorised, regardless of how much work remains.")
line("Budget variance", "TOR-approved budget less spend to date. Negative means overspent.")
line("Approved scope changes", "Cumulative change requests approved. A rising count with a flat "
                               "budget is the classic TOR drift signal - scope grew, funding did "
                               "not.")
line("RACI gaps", "Deliverables with no accountable owner named. Any number above zero is a "
                  "governance finding, not a delivery one.")
line("Calculated RAG", "Computed objectively from the thresholds in Dashboard C7:C11, so it "
                       "cannot be talked up or down. Shown next to the PM's own RAG - a "
                       "disagreement between the two is itself worth discussing.")
gap()

section("Colour legend")
for lbl, col, fil, desc in [
    ("Blue text", BLUE_INPUT, None, "A number or value you type in. Safe to edit."),
    ("Yellow fill", BLACK, YELLOW_FILL, "Key input you are expected to set - cadence, reporting "
                                        "date, RAG thresholds, roll-up date."),
    ("Black text", BLACK, None, "Formula calculated on that sheet. Do not overwrite."),
    ("Green text", GREEN_LINK, None, "Value pulled from another sheet. Do not overwrite."),
]:
    a = ws.cell(row=r, column=2, value=lbl)
    a.font = Font(name=FONT, size=9, bold=True, color=col)
    if fil:
        a.fill = PatternFill("solid", fgColor=fil)
    a.alignment = Alignment(vertical="center")
    b = ws.cell(row=r, column=3, value=desc)
    b.font = Font(name=FONT, size=9)
    r += 1
gap()

section("Capacity and assumptions")
line("Row capacity", f"TOR Register {TOR_LAST - TOR_FIRST + 1} projects; Daily Log "
                     f"{DAY_LAST - DAY_FIRST + 1} rows; Monthly Log {LOG_LAST - LOG_FIRST + 1} "
                     f"rows; Deliverables and Risks {DEL_LAST - DEL_FIRST + 1} rows each. "
                     "Formulas already cover every row in those ranges - just type into the next "
                     "empty row.")
line("Date format", "Daily Log takes the actual date. Monthly Log takes the FIRST DAY of the month "
                    "(e.g. 2026-09-01) - the Dashboard matches on the exact date, so a mid-month "
                    "date will not be picked up in Monthly cadence.")
line("Example data", "The workbook ships with EXAMPLE data for five fictional projects: six months "
                     "on 'Monthly Log' and the working days of the final month on 'Daily Log'. "
                     "Names and figures are invented. Delete rows 4 and below on TOR Register, "
                     "Daily Log, Monthly Log, Deliverables & RACI and Risks & Issues before real "
                     "use.")
line("Currency", "All money is in whole dollars. Change the number formats if the team reports in "
                 "another unit.")
line("Source", "Structure and metric definitions were specified for this request; there is no "
               "external data source behind the example figures.")

for row in ws.iter_rows(min_row=4, max_row=r, min_col=2, max_col=3):
    for c in row:
        if c.alignment.wrap_text is None:
            c.alignment = Alignment(vertical="top", wrap_text=True)

# ========================================================== TOR REGISTER =====
tor = wb.create_sheet("TOR Register")
tor.sheet_properties.tabColor = SLATE
TOR_H = ["Project ID", "Project Name", "Strategic Pillar", "Sponsor",
         "Project Manager", "TOR Approved", "Planned Start", "Planned End",
         "TOR Approved Budget ($)", "Planned Milestones (total)",
         "Planned Deliverables (total)", "TOR Objective",
         "Key In-Scope (per TOR)", "Key Out-of-Scope (per TOR)"]
title_block(tor, "  TOR Register - the approved baseline",
            "  One row per project, straight from the signed Terms of Reference. "
            "Change only via an approved CR. EXAMPLE DATA - replace with the team's projects.",
            len(TOR_H))
header_row(tor, 3, TOR_H)
widths(tor, {"A": 10, "B": 26, "C": 20, "D": 14, "E": 15, "F": 13, "G": 13,
             "H": 13, "I": 18, "J": 13, "K": 13, "L": 46, "M": 42, "N": 38})

for i, p in enumerate(PROJECTS):
    row = TOR_FIRST + i
    for j, v in enumerate(p, start=1):
        c = tor.cell(row=row, column=j, value=v)
        fmt = DATE if j in (6, 7, 8) else (CUR if j == 9 else (INT if j in (10, 11) else None))
        body(c, fmt=fmt, color=BLUE_INPUT, wrap=j >= 12,
             align="center" if j in (1, 6, 7, 8) else None)
    tor.row_dimensions[row].height = 30

for row in range(TOR_FIRST + len(PROJECTS), TOR_LAST + 1):
    for j in range(1, len(TOR_H) + 1):
        fmt = DATE if j in (6, 7, 8) else (CUR if j == 9 else (INT if j in (10, 11) else None))
        body(tor.cell(row=row, column=j), fmt=fmt, color=BLUE_INPUT)
band(tor, TOR_FIRST, TOR_LAST, len(TOR_H))
tor.freeze_panes = "C4"
tor.auto_filter.ref = f"A3:{get_column_letter(len(TOR_H))}{TOR_LAST}"

# ------------------------------------------------ shared log sheet builder ---
LOG_TAIL = ["Milestones Planned to Date", "Milestones Achieved to Date",
            "Deliverables Due to Date", "Deliverables Accepted to Date",
            "Budget Spent to Date ($)", "Change Requests Raised (cum.)",
            "Change Requests Approved (cum.)", "Open Risks (snapshot)",
            "High Risks (snapshot)", "RACI Gaps (snapshot)", "PM RAG",
            "Commentary", "Match Key (auto)"]
LOG_W = {"B": 10, "C": 12, "D": 12, "E": 12, "F": 12, "G": 16, "H": 12, "I": 12,
         "J": 11, "K": 10, "L": 10, "M": 9, "N": 52, "O": 18}


def write_log_sheet(ws_, first, last, date_header, date_fmt, key_fmt, rows):
    """Daily and Monthly logs are the same grid; only the date column differs."""
    headers = [date_header] + LOG_TAIL
    header_row(ws_, 3, headers)
    widths(ws_, {"A": 15, **LOG_W})

    r_ = first
    for vals in rows:
        for j, v in enumerate(vals, start=1):
            c = ws_.cell(row=r_, column=j, value=v)
            fmt = date_fmt if j == 1 else (CUR if j == 7 else (INT if 3 <= j <= 12 else None))
            body(c, fmt=fmt, color=BLUE_INPUT, wrap=(j == 14),
                 align="center" if j in (1, 2, 13) else None)
        r_ += 1

    for rr in range(first, last + 1):
        if rr >= r_:
            for j in range(1, 15):
                fmt = date_fmt if j == 1 else (CUR if j == 7 else (INT if 3 <= j <= 12 else None))
                body(ws_.cell(row=rr, column=j), fmt=fmt, color=BLUE_INPUT,
                     align="center" if j in (1, 2, 13) else None)
        k = ws_.cell(row=rr, column=15,
                     value=f'=IF($B{rr}="","",TEXT($A{rr},"{key_fmt}")&"|"&$B{rr})')
        body(k, color=GREY, align="center")

    band(ws_, first, last, len(headers))
    ws_.freeze_panes = "C4"
    ws_.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{last}"


# ============================================================= DAILY LOG =====
day = wb.create_sheet("Daily Log")
day.sheet_properties.tabColor = "2E75B6"
daily_rows = []
for d_i, dte in enumerate(DAILY_DAYS):
    for pid in [p[0] for p in PROJECTS]:
        s = DAILY[pid]
        daily_rows.append([dte, pid, s["ms_plan"][d_i], s["ms_done"][d_i],
                           s["dl_due"][d_i], s["dl_acc"][d_i], s["spend"][d_i],
                           s["cr_raised"][d_i], s["cr_appr"][d_i], s["risks"][d_i],
                           s["high"][d_i], s["gaps"][d_i], s["rag"][d_i],
                           s["note"][d_i]])
title_block(day, "  Daily Log - day-by-day tracking inside the month",
            "  One row per project per working day. ALL FIGURES ARE CUMULATIVE TO DATE. "
            "EXAMPLE DATA - delete before real use.", 15)
write_log_sheet(day, DAY_FIRST, DAY_LAST, "Date", DATE, "YYYY-MM-DD", daily_rows)

# month-end roll-up: read any date, copy the result into the Monthly Log
day["Q2"] = "Month-end roll-up - set the date, copy the row into the Monthly Log"
day["Q2"].font = Font(name=FONT, size=9, bold=True, italic=True, color=NAVY)
day["Q3"] = "Roll-up date:"
day["Q3"].font = Font(name=FONT, size=9, bold=True, color=NAVY)
day["R3"] = REPORT_DAY
body(day["R3"], fmt=DATE, color=BLUE_INPUT, bold=True, align="center", fill=YELLOW_FILL)

ROLL_H = ["Project", "Ms Planned", "Ms Achieved", "Dl Due", "Dl Accepted",
          "Spend to Date ($)", "CR Raised", "CR Approved", "Open Risks",
          "High Risks", "RACI Gaps", "PM RAG"]
for j, h in enumerate(ROLL_H, start=17):
    c = day.cell(row=4, column=j, value=h)
    c.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=SLATE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BOX
day.row_dimensions[4].height = 30
for j in range(17, 29):
    day.column_dimensions[get_column_letter(j)].width = 13

ROLL_FIRST = 5
ROLL_LAST = ROLL_FIRST + (TOR_LAST - TOR_FIRST)
for i in range(ROLL_LAST - ROLL_FIRST + 1):
    rr = ROLL_FIRST + i
    tor_row = TOR_FIRST + i
    body(day.cell(row=rr, column=17,
                  value=f"=IF('TOR Register'!$A{tor_row}=\"\",\"\",'TOR Register'!$A{tor_row})"),
         color=GREEN_LINK, align="center")
    for off, src in enumerate("CDEFGHIJKLM"):
        col = 18 + off
        fmt = CUR if src == "G" else (INT if src in "CDEFHIJKL" else None)
        body(day.cell(row=rr, column=col,
                      value=f'=IFERROR(INDEX(${src}${DAY_FIRST}:${src}${DAY_LAST},'
                            f'MATCH(TEXT($R$3,"YYYY-MM-DD")&"|"&$Q{rr},'
                            f'$O${DAY_FIRST}:$O${DAY_LAST},0)),"")'),
             fmt=fmt, align="center" if src != "G" else None)

# =========================================================== MONTHLY LOG =====
log = wb.create_sheet("Monthly Log")
log.sheet_properties.tabColor = SLATE
monthly_rows = []
for m_i, month in enumerate(MONTHS):
    for pid in [p[0] for p in PROJECTS]:
        d = LOG[pid]
        monthly_rows.append([month, pid, d["ms_plan"][m_i], d["ms_done"][m_i],
                             d["dl_due"][m_i], d["dl_acc"][m_i], d["spend"][m_i],
                             d["cr_raised"][m_i], d["cr_appr"][m_i], d["risks"][m_i],
                             d["high"][m_i], d["gaps"][m_i], d["rag"][m_i],
                             d["note"][m_i]])
title_block(log, "  Monthly Log - the governance record",
            "  One row per project per month. ALL FIGURES ARE CUMULATIVE TO DATE. Enter the first "
            "day of the month. EXAMPLE DATA - delete before real use.", 15)
write_log_sheet(log, LOG_FIRST, LOG_LAST, "Month (1st of month)", MONTH, "YYYY-MM", monthly_rows)

# ==================================================== DELIVERABLES & RACI =====
dl = wb.create_sheet("Deliverables & RACI")
dl.sheet_properties.tabColor = SLATE
DL_H = ["Deliverable ID", "Project ID", "Deliverable (per TOR)", "TOR Clause Ref",
        "Responsible", "Accountable", "Consulted", "Informed", "Due Date", "Status"]
title_block(dl, "  Deliverables & RACI - TOR outputs and who owns them",
            "  A blank Accountable is a governance gap and is counted automatically. "
            "EXAMPLE DATA - delete before real use.", len(DL_H))
header_row(dl, 3, DL_H)
widths(dl, {"A": 13, "B": 10, "C": 36, "D": 17, "E": 15, "F": 15, "G": 17,
            "H": 15, "I": 12, "J": 13, "K": 3, "L": 10, "M": 13, "N": 13, "O": 13})

for i, d in enumerate(DELIVERABLES):
    r_ = DEL_FIRST + i
    for j, v in enumerate(d, start=1):
        c = dl.cell(row=r_, column=j, value=v)
        body(c, fmt=DATE if j == 9 else None, color=BLUE_INPUT, wrap=(j == 3),
             align="center" if j in (1, 2, 9, 10) else None)
        if j == 6 and v == "":
            c.fill = PatternFill("solid", fgColor="FCE4E4")
for r_ in range(DEL_FIRST + len(DELIVERABLES), DEL_LAST + 1):
    for j in range(1, len(DL_H) + 1):
        body(dl.cell(row=r_, column=j), fmt=DATE if j == 9 else None,
             color=BLUE_INPUT, align="center" if j in (1, 2, 9, 10) else None)
band(dl, DEL_FIRST, DEL_LAST, len(DL_H))

dl["L2"] = "Live register counts"
dl["L2"].font = Font(name=FONT, size=9, bold=True, italic=True, color=NAVY)
for j, h in enumerate(["Project", "In Register", "Accepted",
                       "RACI Gaps -> copy to log"], start=12):
    c = dl.cell(row=3, column=j, value=h)
    c.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=SLATE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BOX
for i, p in enumerate(PROJECTS):
    r_ = 4 + i
    body(dl.cell(row=r_, column=12, value=p[0]), color=GREEN_LINK, align="center")
    body(dl.cell(row=r_, column=13,
                 value=f'=COUNTIFS($B${DEL_FIRST}:$B${DEL_LAST},$L{r_})'),
         fmt=INT, align="center")
    body(dl.cell(row=r_, column=14,
                 value=f'=COUNTIFS($B${DEL_FIRST}:$B${DEL_LAST},$L{r_},'
                       f'$J${DEL_FIRST}:$J${DEL_LAST},"Accepted")'), fmt=INT, align="center")
    body(dl.cell(row=r_, column=15,
                 value=f'=COUNTIFS($B${DEL_FIRST}:$B${DEL_LAST},$L{r_},'
                       f'$F${DEL_FIRST}:$F${DEL_LAST},"")'), fmt=INT, align="center")

note = dl.cell(row=4 + len(PROJECTS) + 1, column=12,
               value="Only the RACI gap count is copied into the logs. This register holds the "
                     "major TOR deliverables and their ownership; the logs' 'due / accepted' "
                     "counts come from the full delivery plan, so they are normally larger.")
note.font = Font(name=FONT, size=8, italic=True, color=GREY)
note.alignment = Alignment(vertical="top", wrap_text=True)
dl.merge_cells(start_row=4 + len(PROJECTS) + 1, start_column=12,
               end_row=4 + len(PROJECTS) + 3, end_column=15)
dl.freeze_panes = "C4"
dl.auto_filter.ref = f"A3:{get_column_letter(len(DL_H))}{DEL_LAST}"

# ========================================================= RISKS & ISSUES =====
rk = wb.create_sheet("Risks & Issues")
rk.sheet_properties.tabColor = SLATE
RK_H = ["ID", "Project ID", "Type", "Description", "Likelihood (1-5)",
        "Impact (1-5)", "Score", "Severity", "Owner", "Mitigation / Action",
        "Status", "Raised", "Target Close"]
title_block(rk, "  Risks & Issues register",
            "  Score = likelihood x impact. Severity and the live counts are calculated. "
            "EXAMPLE DATA - delete before real use.", len(RK_H))
header_row(rk, 3, RK_H)
widths(rk, {"A": 9, "B": 10, "C": 9, "D": 46, "E": 11, "F": 10, "G": 8, "H": 11,
            "I": 14, "J": 46, "K": 12, "L": 12, "M": 12, "N": 3, "O": 10,
            "P": 11, "Q": 11, "R": 11})

for i, k in enumerate(RISKS):
    r_ = RSK_FIRST + i
    rid, pid, typ, desc, lik, imp, owner, mit, status, raised, target = k
    vals = [rid, pid, typ, desc, lik, imp, None, None, owner, mit, status, raised, target]
    for j, v in enumerate(vals, start=1):
        c = rk.cell(row=r_, column=j, value=v)
        body(c, fmt=DATE if j in (12, 13) else None, color=BLUE_INPUT,
             wrap=j in (4, 10), align="center" if j in (1, 2, 3, 5, 6, 11, 12, 13) else None)
for r_ in range(RSK_FIRST, RSK_LAST + 1):
    if r_ >= RSK_FIRST + len(RISKS):
        for j in list(range(1, 7)) + list(range(9, 14)):
            body(rk.cell(row=r_, column=j), fmt=DATE if j in (12, 13) else None,
                 color=BLUE_INPUT, align="center" if j in (1, 2, 3, 5, 6, 11, 12, 13) else None)
    body(rk.cell(row=r_, column=7, value=f'=IF($E{r_}="","",$E{r_}*$F{r_})'),
         fmt=INT, align="center")
    body(rk.cell(row=r_, column=8,
                 value=f'=IF($G{r_}="","",IF($G{r_}>=15,"High",IF($G{r_}>=8,"Medium","Low")))'),
         align="center", bold=True)
band(rk, RSK_FIRST, RSK_LAST, len(RK_H))

rk["O2"] = "Live counts - copy into the logs"
rk["O2"].font = Font(name=FONT, size=9, bold=True, italic=True, color=NAVY)
for j, h in enumerate(["Project", "Open Risks -> copy to log",
                       "High Risks -> copy to log"], start=15):
    c = rk.cell(row=3, column=j, value=h)
    c.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=SLATE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BOX
for i, p in enumerate(PROJECTS):
    r_ = 4 + i
    body(rk.cell(row=r_, column=15, value=p[0]), color=GREEN_LINK, align="center")
    # "Open" for reporting purposes = anything not yet closed.
    body(rk.cell(row=r_, column=16,
                 value=f'=COUNTIFS($B${RSK_FIRST}:$B${RSK_LAST},$O{r_},'
                       f'$K${RSK_FIRST}:$K${RSK_LAST},"Open")'
                       f'+COUNTIFS($B${RSK_FIRST}:$B${RSK_LAST},$O{r_},'
                       f'$K${RSK_FIRST}:$K${RSK_LAST},"Mitigating")'), fmt=INT, align="center")
    body(rk.cell(row=r_, column=17,
                 value=f'=COUNTIFS($B${RSK_FIRST}:$B${RSK_LAST},$O{r_},'
                       f'$H${RSK_FIRST}:$H${RSK_LAST},"High",'
                       f'$K${RSK_FIRST}:$K${RSK_LAST},"<>Closed")'), fmt=INT, align="center")

for sev, fill in SEVERITY_FILL.items():
    rk.conditional_formatting.add(
        f"H{RSK_FIRST}:H{RSK_LAST}",
        CellIsRule(operator="equal", formula=[f'"{sev}"'], fill=fill))
rk.freeze_panes = "C4"
rk.auto_filter.ref = f"A3:{get_column_letter(len(RK_H))}{RSK_LAST}"

# ============================================================== DASHBOARD =====
db = wb.create_sheet("Dashboard")
db.sheet_properties.tabColor = "C00000"
db.sheet_view.showGridLines = False
DB_H = ["Project ID", "Project Name", "PM", "TOR Budget ($)", "Spent to Date ($)",
        "Budget Used %", "Budget Variance ($)", "Milestones Planned",
        "Milestones Achieved", "Milestone Hit Rate %", "Deliverables Due",
        "Deliverables Accepted", "Scope Adherence %", "Approved Scope Changes",
        "Open Risks", "High Risks", "RACI Gaps", "PM RAG", "Calculated RAG",
        "PM Commentary"]
title_block(db, "  Dashboard - delivery against TOR",
            "  Set the cadence in C3 and the reporting date in C4. Everything below "
            "recalculates. Red and Amber rows are the agenda.", len(DB_H))
widths(db, {"A": 10, "B": 25, "C": 14, "D": 14, "E": 15, "F": 12, "G": 15,
            "J": 13, "M": 14, "N": 13, "O": 10, "P": 10, "Q": 10, "R": 10,
            "S": 13, "T": 56})
for col in "HIKL":
    db.column_dimensions[col].width = 11

# --- controls
db["B3"] = "Cadence:"
db["B3"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
db["C3"] = "Monthly"
body(db["C3"], align="center", fill=YELLOW_FILL)
db["C3"].font = Font(name=FONT, size=10, bold=True, color=BLUE_INPUT)
db["D3"] = "<- Daily reads the Daily Log; Monthly reads the Monthly Log"
db["D3"].font = Font(name=FONT, size=9, italic=True, color=GREY)

db["B4"] = "Reporting date:"
db["B4"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
db["C4"] = REPORT_MONTH
body(db["C4"], fmt=DATE, align="center", fill=YELLOW_FILL)
db["C4"].font = Font(name=FONT, size=10, bold=True, color=BLUE_INPUT)
db["D4"] = "<- Daily: the actual date. Monthly: the FIRST DAY of the month, e.g. 2026-09-01"
db["D4"].font = Font(name=FONT, size=9, italic=True, color=GREY)

db["B6"] = "RAG thresholds (edit to match the team's tolerance)"
db["B6"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
THRESHOLDS = [
    ("Milestone hit rate - Amber below", 0.90, PCT,
     "Below this, schedule is off the TOR baseline enough to flag."),
    ("Milestone hit rate - Red below", 0.75, PCT,
     "Below this, the TOR schedule is not credible without a re-baseline."),
    ("Budget used - Amber above", 0.95, PCT,
     "Approaching the TOR-authorised budget."),
    ("Budget used - Red above", 1.00, PCT,
     "Spent past what the TOR authorised."),
    ("High risks - Red at or above", 2, INT,
     "This many unclosed high-severity risks forces a Red."),
]
for i, (lbl, val, fmt, note_txt) in enumerate(THRESHOLDS):
    r_ = 7 + i
    a = db.cell(row=r_, column=2, value=lbl)
    a.font = Font(name=FONT, size=9, color=SLATE)
    c = db.cell(row=r_, column=3, value=val)
    body(c, fmt=fmt, color=BLUE_INPUT, bold=True, align="center", fill=YELLOW_FILL)
    n = db.cell(row=r_, column=4, value=note_txt)
    n.font = Font(name=FONT, size=8, italic=True, color=GREY)
T_MS_AMBER, T_MS_RED, T_BU_AMBER, T_BU_RED, T_HR_RED = ("$C$7", "$C$8", "$C$9", "$C$10", "$C$11")

db["G6"] = "Portfolio for the selected cadence and date"
db["G6"].font = Font(name=FONT, size=10, bold=True, color=NAVY)

PROW_FIRST = 16
PROW_LAST = PROW_FIRST + (TOR_LAST - TOR_FIRST)


def cad_sum(col_letter, r_):
    """SUMIFS against whichever log the cadence cell selects."""
    return (f'=IF($C$3="Daily",'
            f"SUMIFS('Daily Log'!${col_letter}${DAY_FIRST}:${col_letter}${DAY_LAST},"
            f"'Daily Log'!$A${DAY_FIRST}:$A${DAY_LAST},$C$4,"
            f"'Daily Log'!$B${DAY_FIRST}:$B${DAY_LAST},$A{r_}),"
            f"SUMIFS('Monthly Log'!${col_letter}${LOG_FIRST}:${col_letter}${LOG_LAST},"
            f"'Monthly Log'!$A${LOG_FIRST}:$A${LOG_LAST},$C$4,"
            f"'Monthly Log'!$B${LOG_FIRST}:$B${LOG_LAST},$A{r_}))")


def cad_text(col_letter, r_):
    """Text lookup via each log's match key; the key granularity differs."""
    return (f'=IFERROR(IF($C$3="Daily",'
            f"INDEX('Daily Log'!${col_letter}${DAY_FIRST}:${col_letter}${DAY_LAST},"
            f'MATCH(TEXT($C$4,"YYYY-MM-DD")&"|"&$A{r_},'
            f"'Daily Log'!$O${DAY_FIRST}:$O${DAY_LAST},0)),"
            f"INDEX('Monthly Log'!${col_letter}${LOG_FIRST}:${col_letter}${LOG_LAST},"
            f'MATCH(TEXT($C$4,"YYYY-MM")&"|"&$A{r_},'
            f"'Monthly Log'!$O${LOG_FIRST}:$O${LOG_LAST},0))),\"\")")


TILES = [
    ("Projects reporting",
     f'=IF($C$3="Daily",COUNTIFS(\'Daily Log\'!$A${DAY_FIRST}:$A${DAY_LAST},$C$4),'
     f'COUNTIFS(\'Monthly Log\'!$A${LOG_FIRST}:$A${LOG_LAST},$C$4))', INT),
    ("TOR budget ($)",
     f'=SUMIF($A${PROW_FIRST}:$A${PROW_LAST},"<>",$D${PROW_FIRST}:$D${PROW_LAST})', CUR),
    ("Spent to date ($)", f'=SUM($E${PROW_FIRST}:$E${PROW_LAST})', CUR),
    ("Budget used %", f'=IF(SUM($D${PROW_FIRST}:$D${PROW_LAST})=0,"",'
                      f'SUM($E${PROW_FIRST}:$E${PROW_LAST})/SUM($D${PROW_FIRST}:$D${PROW_LAST}))', PCT),
    ("Milestone hit rate %", f'=IF(SUM($H${PROW_FIRST}:$H${PROW_LAST})=0,"",'
                             f'SUM($I${PROW_FIRST}:$I${PROW_LAST})/SUM($H${PROW_FIRST}:$H${PROW_LAST}))', PCT),
    ("Scope adherence %", f'=IF(SUM($K${PROW_FIRST}:$K${PROW_LAST})=0,"",'
                          f'SUM($L${PROW_FIRST}:$L${PROW_LAST})/SUM($K${PROW_FIRST}:$K${PROW_LAST}))', PCT),
    ("Open risks", f'=SUM($O${PROW_FIRST}:$O${PROW_LAST})', INT),
    ("High risks", f'=SUM($P${PROW_FIRST}:$P${PROW_LAST})', INT),
    ("RACI gaps", f'=SUM($Q${PROW_FIRST}:$Q${PROW_LAST})', INT),
    ("Projects Red", f'=COUNTIF($S${PROW_FIRST}:$S${PROW_LAST},"Red")', INT),
    ("Projects Amber", f'=COUNTIF($S${PROW_FIRST}:$S${PROW_LAST},"Amber")', INT),
    ("Projects Green", f'=COUNTIF($S${PROW_FIRST}:$S${PROW_LAST},"Green")', INT),
]
for i, (lbl, formula, fmt) in enumerate(TILES):
    col = 7 + (i % 6) * 2          # G, I, K, M, O, Q
    r_ = 7 if i < 6 else 10
    a = db.cell(row=r_, column=col, value=lbl)
    a.font = Font(name=FONT, size=8, bold=True, color="FFFFFF")
    a.fill = PatternFill("solid", fgColor=SLATE)
    a.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    a.border = BOX
    db.merge_cells(start_row=r_, start_column=col, end_row=r_, end_column=col + 1)
    v = db.cell(row=r_ + 1, column=col, value=formula)
    v.font = Font(name=FONT, size=12, bold=True, color=NAVY)
    v.number_format = fmt
    v.alignment = Alignment(horizontal="center", vertical="center")
    v.fill = PatternFill("solid", fgColor=LIGHT)
    v.border = BOX
    db.merge_cells(start_row=r_ + 1, start_column=col, end_row=r_ + 1, end_column=col + 1)
    db.row_dimensions[r_].height = 24
    db.row_dimensions[r_ + 1].height = 22

db["A14"] = "Project detail for the selected cadence and date"
db["A14"].font = Font(name=FONT, size=11, bold=True, color=NAVY)
header_row(db, 15, DB_H)

L_ID = f"'TOR Register'!$A${TOR_FIRST}:$A${TOR_LAST}"


def tor_lookup(col_letter, r_):
    return (f"=IFERROR(INDEX('TOR Register'!${col_letter}${TOR_FIRST}:${col_letter}${TOR_LAST},"
            f"MATCH($A{r_},{L_ID},0)),\"\")")


for i in range(PROW_LAST - PROW_FIRST + 1):
    r_ = PROW_FIRST + i
    tor_row = TOR_FIRST + i
    has_data = i < len(PROJECTS)

    body(db.cell(row=r_, column=1, value=f"=IF('TOR Register'!$A{tor_row}=\"\",\"\","
                                         f"'TOR Register'!$A{tor_row})"),
         color=GREEN_LINK, align="center", bold=True)
    body(db.cell(row=r_, column=2, value=tor_lookup("B", r_)), color=GREEN_LINK)
    body(db.cell(row=r_, column=3, value=tor_lookup("E", r_)), color=GREEN_LINK)
    body(db.cell(row=r_, column=4, value=tor_lookup("I", r_)), fmt=CUR, color=GREEN_LINK)
    body(db.cell(row=r_, column=5, value=cad_sum("G", r_)), fmt=CUR, color=GREEN_LINK)
    body(db.cell(row=r_, column=6,
                 value=f'=IF(N($D{r_})=0,"",$E{r_}/$D{r_})'), fmt=PCT, align="center")
    body(db.cell(row=r_, column=7,
                 value=f'=IF($A{r_}="","",$D{r_}-$E{r_})'), fmt=CUR)
    body(db.cell(row=r_, column=8, value=cad_sum("C", r_)), fmt=INT, color=GREEN_LINK, align="center")
    body(db.cell(row=r_, column=9, value=cad_sum("D", r_)), fmt=INT, color=GREEN_LINK, align="center")
    body(db.cell(row=r_, column=10,
                 value=f'=IF($H{r_}=0,"",$I{r_}/$H{r_})'), fmt=PCT, align="center", bold=True)
    body(db.cell(row=r_, column=11, value=cad_sum("E", r_)), fmt=INT, color=GREEN_LINK, align="center")
    body(db.cell(row=r_, column=12, value=cad_sum("F", r_)), fmt=INT, color=GREEN_LINK, align="center")
    body(db.cell(row=r_, column=13,
                 value=f'=IF($K{r_}=0,"",$L{r_}/$K{r_})'), fmt=PCT, align="center", bold=True)
    body(db.cell(row=r_, column=14, value=cad_sum("I", r_)), fmt=INT, color=GREEN_LINK, align="center")
    body(db.cell(row=r_, column=15, value=cad_sum("J", r_)), fmt=INT, color=GREEN_LINK, align="center")
    body(db.cell(row=r_, column=16, value=cad_sum("K", r_)), fmt=INT, color=GREEN_LINK, align="center")
    body(db.cell(row=r_, column=17, value=cad_sum("L", r_)), fmt=INT, color=GREEN_LINK, align="center")
    body(db.cell(row=r_, column=18, value=cad_text("M", r_)), color=GREEN_LINK,
         align="center", bold=True)

    # Calculated RAG - recomputed from raw counts, never from the display rates,
    # so an empty "nothing due yet" cell cannot read as 0%.
    rag = (
        f'=IF($A{r_}="","",'
        f'IF(AND($H{r_}=0,$K{r_}=0,$E{r_}=0),"Not started",'
        f'IF(OR('
        f'AND($D{r_}>0,$E{r_}/$D{r_}>{T_BU_RED}),'
        f'AND($H{r_}>0,$I{r_}/$H{r_}<{T_MS_RED}),'
        f'$P{r_}>={T_HR_RED}'
        f'),"Red",'
        f'IF(OR('
        f'AND($D{r_}>0,$E{r_}/$D{r_}>{T_BU_AMBER}),'
        f'AND($H{r_}>0,$I{r_}/$H{r_}<{T_MS_AMBER}),'
        f'AND($K{r_}>0,$L{r_}/$K{r_}<{T_MS_AMBER}),'
        f'$P{r_}>0,'
        f'$Q{r_}>0'
        f'),"Amber","Green"))))'
    )
    body(db.cell(row=r_, column=19, value=rag), align="center", bold=True)
    body(db.cell(row=r_, column=20, value=cad_text("N", r_)), color=GREEN_LINK, wrap=True)
    db.row_dimensions[r_].height = 26 if has_data else 15

for col in ("R", "S"):
    for label, fill in RAG_FILL.items():
        db.conditional_formatting.add(
            f"{col}{PROW_FIRST}:{col}{PROW_LAST}",
            CellIsRule(operator="equal", formula=[f'"{label}"'], fill=fill))
db.freeze_panes = "D16"

fn = PROW_LAST + 2
db.cell(row=fn, column=1, value="Notes and assumptions").font = Font(
    name=FONT, size=10, bold=True, color=NAVY)
for i, t in enumerate([
    "Cadence C3 chooses the source: Daily reads 'Daily Log' and matches the exact date in C4; "
    "Monthly reads 'Monthly Log' and matches the first day of the month in C4.",
    "Both logs hold cumulative-to-date figures, so every metric here is 'position as at the "
    "reporting date', not activity within that day or month.",
    "Because both logs are cumulative, the last daily row of a month equals that month's monthly "
    "row - the two cadences agree at month end by construction.",
    "Scope adherence = deliverables accepted / deliverables due to date. Approved scope changes "
    "are shown separately: a rising CR count against a flat TOR budget is scope drift, even when "
    "the adherence percentage looks healthy.",
    "Calculated RAG is derived only from the thresholds in C7:C11 and the raw counts, so it is "
    "reproducible. Where it disagrees with the PM RAG, the difference is the discussion.",
    "'Not started' appears when a project has no milestones due, no deliverables due and no spend "
    "as at the reporting date.",
    "Rows are driven by the TOR Register: add a project there and it appears here automatically, "
    "up to the register's capacity.",
    "Open risks counts anything not Closed (Open + Mitigating) on the Risks & Issues register at "
    "the point the snapshot was taken.",
], start=0):
    c = db.cell(row=fn + 1 + i, column=1, value=f"- {t}")
    c.font = Font(name=FONT, size=8, italic=True, color=GREY)
    c.alignment = Alignment(vertical="top", wrap_text=True)
    db.merge_cells(start_row=fn + 1 + i, start_column=1, end_row=fn + 1 + i, end_column=13)
    db.row_dimensions[fn + 1 + i].height = 22

# ================================================================= CHARTS =====
ch = wb.create_sheet("Charts")
ch.sheet_properties.tabColor = "2E75B6"
ch.sheet_view.showGridLines = False
title_block(ch, "  Trend and comparison charts",
            "  Monthly trend and per-project comparison at the top; daily trend for the current "
            "month below. Add a row to a block and its chart extends.", 12)

TREND_H = ["Month", "Milestones Planned", "Milestones Achieved", "Milestone Hit Rate %",
           "Deliverables Due", "Deliverables Accepted", "Scope Adherence %",
           "Budget Spent to Date ($)", "Open Risks", "High Risks", "RACI Gaps"]
header_row(ch, 4, TREND_H)
widths(ch, {"A": 13, "B": 12, "C": 12, "D": 12, "E": 12, "F": 12, "G": 12,
            "H": 17, "I": 10, "J": 10, "K": 10})

TR_FIRST = 5
TR_LAST = TR_FIRST + 23          # 24 months of capacity


def period_sum(sheet_name, col_letter, rr, first, last):
    return (f"=IF($A{rr}=\"\",\"\",SUMIFS('{sheet_name}'!${col_letter}${first}:${col_letter}${last},"
            f"'{sheet_name}'!$A${first}:$A${last},$A{rr}))")


for i in range(TR_LAST - TR_FIRST + 1):
    r_ = TR_FIRST + i
    val = MONTHS[i] if i < len(MONTHS) else None
    body(ch.cell(row=r_, column=1, value=val), fmt=MONTH, color=BLUE_INPUT, align="center")
    for col, src in ((2, "C"), (3, "D"), (5, "E"), (6, "F"), (9, "J"), (10, "K"), (11, "L")):
        body(ch.cell(row=r_, column=col,
                     value=period_sum("Monthly Log", src, r_, LOG_FIRST, LOG_LAST)),
             fmt=INT, align="center")
    body(ch.cell(row=r_, column=4, value=f'=IF(N($B{r_})=0,"",$C{r_}/$B{r_})'),
         fmt=PCT, align="center")
    body(ch.cell(row=r_, column=7, value=f'=IF(N($E{r_})=0,"",$F{r_}/$E{r_})'),
         fmt=PCT, align="center")
    body(ch.cell(row=r_, column=8,
                 value=period_sum("Monthly Log", "G", r_, LOG_FIRST, LOG_LAST)), fmt=CUR)
band(ch, TR_FIRST, TR_LAST, len(TREND_H))

# per-project comparison, mirroring the Dashboard for the selected cadence/date
CMP_FIRST = TR_LAST + 3
ch.cell(row=CMP_FIRST - 1, column=1,
        value="Per-project comparison - follows the cadence and date set on the Dashboard "
              "(covers the projects present when this workbook was generated; re-run the "
              "generator after adding projects)").font = Font(
    name=FONT, size=10, bold=True, color=NAVY)
header_row(ch, CMP_FIRST, ["Project", "TOR Budget ($)", "Spent to Date ($)",
                           "Milestone Hit Rate %", "Scope Adherence %"])
for i in range(len(PROJECTS)):
    r_ = CMP_FIRST + 1 + i
    d_ = PROW_FIRST + i
    body(ch.cell(row=r_, column=1, value=f"=IF(Dashboard!$B{d_}=\"\",\"\",Dashboard!$B{d_})"),
         color=GREEN_LINK)
    body(ch.cell(row=r_, column=2, value=f"=Dashboard!$D{d_}"), fmt=CUR, color=GREEN_LINK)
    body(ch.cell(row=r_, column=3, value=f"=Dashboard!$E{d_}"), fmt=CUR, color=GREEN_LINK)
    body(ch.cell(row=r_, column=4, value=f'=IF(Dashboard!$J{d_}="","",Dashboard!$J{d_})'),
         fmt=PCT, color=GREEN_LINK, align="center")
    body(ch.cell(row=r_, column=5, value=f'=IF(Dashboard!$M{d_}="","",Dashboard!$M{d_})'),
         fmt=PCT, color=GREEN_LINK, align="center")
CMP_LAST = CMP_FIRST + len(PROJECTS)

# daily trend block
DAY_TR_HEAD = CMP_LAST + 3
ch.cell(row=DAY_TR_HEAD - 1, column=1,
        value="Daily trend - portfolio position by day, straight from the Daily Log").font = Font(
    name=FONT, size=10, bold=True, color=NAVY)
header_row(ch, DAY_TR_HEAD,
           ["Date", "Milestones Planned", "Milestones Achieved", "Milestone Hit Rate %",
            "Deliverables Due", "Deliverables Accepted", "Scope Adherence %",
            "Budget Spent to Date ($)", "Open Risks", "High Risks", "RACI Gaps"])
DTR_FIRST = DAY_TR_HEAD + 1
DTR_LAST = DTR_FIRST + 64        # 65 days of capacity
for i in range(DTR_LAST - DTR_FIRST + 1):
    r_ = DTR_FIRST + i
    val = DAILY_DAYS[i] if i < len(DAILY_DAYS) else None
    body(ch.cell(row=r_, column=1, value=val), fmt=DATE, color=BLUE_INPUT, align="center")
    for col, src in ((2, "C"), (3, "D"), (5, "E"), (6, "F"), (9, "J"), (10, "K"), (11, "L")):
        body(ch.cell(row=r_, column=col,
                     value=period_sum("Daily Log", src, r_, DAY_FIRST, DAY_LAST)),
             fmt=INT, align="center")
    body(ch.cell(row=r_, column=4, value=f'=IF(N($B{r_})=0,"",$C{r_}/$B{r_})'),
         fmt=PCT, align="center")
    body(ch.cell(row=r_, column=7, value=f'=IF(N($E{r_})=0,"",$F{r_}/$E{r_})'),
         fmt=PCT, align="center")
    body(ch.cell(row=r_, column=8,
                 value=period_sum("Daily Log", "G", r_, DAY_FIRST, DAY_LAST)), fmt=CUR)
band(ch, DTR_FIRST, DTR_LAST, 11)

TR_END = TR_FIRST + len(MONTHS) - 1
DTR_END = DTR_FIRST + len(DAILY_DAYS) - 1


def style_chart(c, title, y_title, h=8.5, w=17):
    c.title = title
    c.y_axis.title = y_title
    c.x_axis.title = None
    c.height, c.width = h, w
    c.style = 2


def line_chart(title, y_title, cols, head_row, first, last, anchor, pct=False):
    c = LineChart()
    style_chart(c, title, y_title)
    if isinstance(cols, tuple):
        c.add_data(Reference(ch, min_col=cols[0], min_row=head_row, max_col=cols[1],
                             max_row=last), titles_from_data=True)
    else:
        for col in cols:
            c.add_data(Reference(ch, min_col=col, min_row=head_row, max_row=last),
                       titles_from_data=True)
    c.set_categories(Reference(ch, min_col=1, min_row=first, max_row=last))
    if pct:
        c.y_axis.numFmt = '0%'
    ch.add_chart(c, anchor)


line_chart("Portfolio delivery against TOR - monthly", "Percent", [4, 7], 4,
           TR_FIRST, TR_END, "M4", pct=True)

c2 = BarChart()
style_chart(c2, "Portfolio spend to date - monthly", "$")
c2.type, c2.grouping = "col", "clustered"
c2.add_data(Reference(ch, min_col=8, min_row=4, max_row=TR_END), titles_from_data=True)
c2.set_categories(Reference(ch, min_col=1, min_row=TR_FIRST, max_row=TR_END))
ch.add_chart(c2, "M22")

line_chart("Risk and governance posture - monthly", "Count", (9, 11), 4,
           TR_FIRST, TR_END, "M40")

c4 = BarChart()
style_chart(c4, "TOR budget vs spend to date, by project", "$")
c4.type, c4.grouping = "col", "clustered"
c4.add_data(Reference(ch, min_col=2, min_row=CMP_FIRST, max_col=3, max_row=CMP_LAST),
            titles_from_data=True)
c4.set_categories(Reference(ch, min_col=1, min_row=CMP_FIRST + 1, max_row=CMP_LAST))
ch.add_chart(c4, "M58")

c5 = BarChart()
style_chart(c5, "Milestone hit rate and scope adherence, by project", "Percent")
c5.type, c5.grouping = "col", "clustered"
c5.add_data(Reference(ch, min_col=4, min_row=CMP_FIRST, max_col=5, max_row=CMP_LAST),
            titles_from_data=True)
c5.set_categories(Reference(ch, min_col=1, min_row=CMP_FIRST + 1, max_row=CMP_LAST))
c5.y_axis.numFmt = '0%'
ch.add_chart(c5, "M76")

line_chart("Portfolio delivery against TOR - daily", "Percent", [4, 7], DAY_TR_HEAD,
           DTR_FIRST, DTR_END, "M94", pct=True)
line_chart("Portfolio spend to date - daily", "$", [8], DAY_TR_HEAD,
           DTR_FIRST, DTR_END, "M112")
line_chart("Risk and governance posture - daily", "Count", (9, 11), DAY_TR_HEAD,
           DTR_FIRST, DTR_END, "M130")

# ============================================================== CEO BRIEF =====
# Five live sentences written off the Dashboard, so the brief follows whatever
# cadence and reporting date the team has selected. Nothing here is typed text.
brief = wb.create_sheet("CEO Brief")
brief.sheet_properties.tabColor = "C55A11"
brief.sheet_view.showGridLines = False
widths(brief, {"A": 3, "B": 5, "C": 22, "D": 22, "E": 22, "F": 22, "G": 22, "H": 22,
               "I": 3, "J": 28, "K": 15, "L": 9, "M": 12, "N": 9, "O": 10, "P": 12})
title_block(brief, "  CEO Brief - the five things to know",
            "  Written from the Dashboard, so it follows the cadence and reporting date set "
            "there. Every sentence below is a live formula, not typed text.", 8)

brief["B4"] = "Cadence:"
brief["B4"].font = Font(name=FONT, size=9, bold=True, color=SLATE)
brief["C4"] = "=Dashboard!$C$3"
body(brief["C4"], color=GREEN_LINK, bold=True)
brief["B5"] = "As at:"
brief["B5"].font = Font(name=FONT, size=9, bold=True, color=SLATE)
brief["C5"] = "=Dashboard!$C$4"
body(brief["C5"], fmt=DATE, color=GREEN_LINK, bold=True)
brief["D5"] = "Change the cadence or date on the Dashboard and this whole page rewrites itself."
brief["D5"].font = Font(name=FONT, size=8, italic=True, color=GREY)

# --- headline strip
STRIP = [
    ("TOR budget", "=Dashboard!$I$8", CUR),
    ("Spent to date", "=Dashboard!$K$8", CUR),
    ("Budget used", "=Dashboard!$M$8", PCT),
    ("Milestone hit rate", "=Dashboard!$O$8", PCT),
    ("Scope adherence", "=Dashboard!$Q$8", PCT),
    ("Red / Amber / Green",
     '=Dashboard!$M$11&"  /  "&Dashboard!$O$11&"  /  "&Dashboard!$Q$11', None),
]
for i, (lbl, formula, fmt) in enumerate(STRIP):
    col = 3 + i
    a = brief.cell(row=7, column=col, value=lbl)
    a.font = Font(name=FONT, size=8, bold=True, color="FFFFFF")
    a.fill = PatternFill("solid", fgColor=SLATE)
    a.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    a.border = BOX
    v = brief.cell(row=8, column=col, value=formula)
    v.font = Font(name=FONT, size=13, bold=True, color=NAVY)
    if fmt:
        v.number_format = fmt
    v.alignment = Alignment(horizontal="center", vertical="center")
    v.fill = PatternFill("solid", fgColor=LIGHT)
    v.border = BOX
brief.row_dimensions[7].height = 24
brief.row_dimensions[8].height = 28

brief["B10"] = "The five things the CEO must know"
brief["B10"].font = Font(name=FONT, size=12, bold=True, color=NAVY)
brief.merge_cells(start_row=10, start_column=2, end_row=10, end_column=8)

# --- the five bullets, each a headline formula plus an explanation formula
H = "Dashboard!"        # portfolio tiles live on the Dashboard
BULLETS = [
    # 1. money
    ('="$"&TEXT(Dashboard!$K$8,"#,##0")&" of the $"&TEXT(Dashboard!$I$8,"#,##0")'
     '&" the TOR authorised is spent - "&TEXT(Dashboard!$M$8,"0%")&" of approved budget."',
     '=IF($K$46=0,'
     '"Every project is still inside the budget its Terms of Reference approved, with $"'
     '&TEXT($K$62,"#,##0")&" of headroom left across the portfolio. No funding decision is '
     'needed this cycle.",'
     '$K$46&" project"&IF($K$46=1,"","s")&IF($K$46=1," has"," have")&" spent past the budget '
     'the TOR authorised, by $"&TEXT($K$47,"#,##0")&" in total. That money is already '
     'committed, so the choice is to fund it, cut scope, or formally re-baseline the TOR - it '
     'is not a forecast that can be left to settle. Headroom elsewhere in the portfolio is $"'
     '&TEXT($K$62,"#,##0")&".")'),
    # 2. delivery
    ('=$K$64&" of "&$K$63&" milestones hit ("&TEXT(Dashboard!$O$8,"0%")&"), and "&$K$66&" of "'
     '&$K$65&" deliverables accepted ("&TEXT(Dashboard!$Q$8,"0%")&")."',
     '="The portfolio is "&$K$59&" milestone"&IF($K$59=1,"","s")&" behind the TOR baseline and "'
     '&$K$60&" deliverable"&IF($K$60=1,"","s")&" short of acceptance"'
     '&IF($K$57="",""," - against a milestone hit rate of "&TEXT($K$57,"0%")'
     '&" the month before")'
     '&". Acceptance is the harder of the two numbers: work can be finished and still not '
     'signed off, and only an accepted deliverable counts as the TOR being met."'),
    # 3. concentration
    ('=IF($K$49="","No project is reporting for this cadence and date.",'
     '$K$49&" is the single largest exposure in the portfolio.")',
     '=IF($K$49="","Set a cadence and reporting date on the Dashboard that has data behind it - '
     'in Monthly cadence the date must be the first day of the month.",'
     '"It is rated "&$K$53&", has hit "&IF($K$50="","no milestones yet",TEXT($K$50,"0%")'
     '&" of its milestones")&", has used "&IF($K$51="","n/a",TEXT($K$51,"0%"))'
     '&" of its approved budget, and carries "&$K$52&" high-severity risk"'
     '&IF($K$52=1,"","s")&". Concentration matters more than the portfolio average: moving this '
     'one project moves the headline numbers more than anything else on the list.")'),
    # 4. scope drift
    ('=$K$61&" change request"&IF($K$61=1,"","s")&" approved across "&$K$48&" project"'
     '&IF($K$48=1,"","s")&", against a TOR budget that has not moved."',
     '=IF($K$61=0,'
     '"Nothing has been added to scope this cycle, so the Terms of Reference still describe what '
     'is actually being built. That is the state worth protecting.",'
     '"Approved scope changes are the quiet way a portfolio outgrows its funding: each one '
     'commits the team to deliver more, but the approved budget on the TOR Register only moves '
     'through a formal re-baseline. The question for the board is whether "&$K$61&" approved '
     'change"&IF($K$61=1,"","s")&IF($K$61=1," was"," were")&" funded - if not, the dates and the '
     'budget in the TOR are both already out of date.")'),
    # 5. governance
    ('=Dashboard!$K$11&" TOR deliverable"&IF(Dashboard!$K$11=1,"","s")'
     '&IF(Dashboard!$K$11=1," has"," have")&" no accountable owner; "&Dashboard!$I$11'
     '&" high-severity risk"&IF(Dashboard!$I$11=1,"","s")&IF(Dashboard!$I$11=1," is"," are")'
     '&" still open."',
     '=IF(Dashboard!$K$11=0,'
     '"Every TOR deliverable currently has a named accountable owner, which is the healthy state '
     'and worth keeping. The "&Dashboard!$G$11&" open risk"&IF(Dashboard!$G$11=1,"","s")'
     '&" on the register stay the thing to watch, "&Dashboard!$I$11&" of them high severity.",'
     '"A deliverable with nobody accountable is a governance finding rather than a delivery one: '
     'when it slips, there is no one to ask. With "&Dashboard!$G$11&" open risk"'
     '&IF(Dashboard!$G$11=1,"","s")&" on the register as well, this is the part of the picture '
     'that will not resolve itself - it needs names attached at the next steering meeting.")'),
]
row = 12
for i, (headline, detail) in enumerate(BULLETS, start=1):
    n = brief.cell(row=row, column=2, value=i)
    n.font = Font(name=FONT, size=13, bold=True, color="FFFFFF")
    n.fill = PatternFill("solid", fgColor=NAVY)
    n.alignment = Alignment(horizontal="center", vertical="center")

    h = brief.cell(row=row, column=3, value=headline)
    h.font = Font(name=FONT, size=11, bold=True, color=NAVY)
    h.alignment = Alignment(vertical="center", wrap_text=True)
    brief.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
    brief.row_dimensions[row].height = 30

    d = brief.cell(row=row + 1, column=3, value=detail)
    d.font = Font(name=FONT, size=9, color="333333")
    d.alignment = Alignment(vertical="top", wrap_text=True)
    brief.merge_cells(start_row=row + 1, start_column=3, end_row=row + 1, end_column=8)
    brief.row_dimensions[row + 1].height = 46
    brief.row_dimensions[row + 2].height = 8
    row += 3

foot = brief.cell(row=row + 1, column=3,
                  value="How to read this: every sentence above is calculated from the Dashboard "
                        "at the cadence and date shown, so it is reproducible and cannot be "
                        "edited into a better story. Bullet 3 ranks projects by an exposure score "
                        "(over-budget 100, each high-severity risk 15, each unowned deliverable 5, "
                        "plus 40 x the share of milestones missed) - a heuristic for 'where to "
                        "look first', not a formal risk measure. The working is in the helper "
                        "block to the right.")
foot.font = Font(name=FONT, size=8, italic=True, color=GREY)
foot.alignment = Alignment(vertical="top", wrap_text=True)
brief.merge_cells(start_row=row + 1, start_column=3, end_row=row + 2, end_column=8)

# --- helper block: per-project working, then the scalars the bullets quote
brief["J3"] = "Calculation helpers - these feed the bullets; safe to ignore"
brief["J3"].font = Font(name=FONT, size=9, bold=True, italic=True, color=NAVY)
for j, h in enumerate(["Project", "Calc RAG", "Over budget", "Overspend $",
                       "Has appr. CR", "Hit rate", "Exposure score"], start=10):
    c = brief.cell(row=4, column=j, value=h)
    c.font = Font(name=FONT, size=8, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=SLATE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BOX
brief.row_dimensions[4].height = 26

HLP_FIRST = 5
HLP_LAST = HLP_FIRST + (TOR_LAST - TOR_FIRST)
for i in range(HLP_LAST - HLP_FIRST + 1):
    hr = HLP_FIRST + i
    dr = PROW_FIRST + i
    body(brief.cell(row=hr, column=10,
                    value=f'=IF(Dashboard!$B{dr}="","",Dashboard!$B{dr})'), color=GREEN_LINK)
    body(brief.cell(row=hr, column=11, value=f'=Dashboard!$S{dr}'), color=GREEN_LINK,
         align="center")
    body(brief.cell(row=hr, column=12,
                    value=f'=IF($J{hr}="",0,IF(AND(Dashboard!$D{dr}>0,'
                          f'Dashboard!$E{dr}>Dashboard!$D{dr}),1,0))'), fmt=INT, align="center")
    body(brief.cell(row=hr, column=13,
                    value=f'=IF($L{hr}=1,Dashboard!$E{dr}-Dashboard!$D{dr},0)'), fmt=CUR)
    body(brief.cell(row=hr, column=14,
                    value=f'=IF($J{hr}="",0,IF(Dashboard!$N{dr}>0,1,0))'), fmt=INT, align="center")
    body(brief.cell(row=hr, column=15,
                    value=f'=IF($J{hr}="","",IF(Dashboard!$H{dr}=0,"",'
                          f'Dashboard!$I{dr}/Dashboard!$H{dr}))'), fmt=PCT, align="center")
    body(brief.cell(row=hr, column=16,
                    value=f'=IF($J{hr}="",-1,$L{hr}*100+Dashboard!$P{dr}*15+Dashboard!$Q{dr}*5'
                          f'+IF(Dashboard!$H{dr}>0,(1-Dashboard!$I{dr}/Dashboard!$H{dr})*40,0))'),
         fmt='0.0', align="center")
band(brief, HLP_FIRST, HLP_LAST, 16)

WORST = f'MATCH(MAX($P${HLP_FIRST}:$P${HLP_LAST}),$P${HLP_FIRST}:$P${HLP_LAST},0)'
ML = f"'Monthly Log'!"
SCALARS = [
    ("Projects over budget", f'=SUM($L${HLP_FIRST}:$L${HLP_LAST})', INT),
    ("Total overspend", f'=SUM($M${HLP_FIRST}:$M${HLP_LAST})', CUR),
    ("Projects with approved CRs", f'=SUM($N${HLP_FIRST}:$N${HLP_LAST})', INT),
    ("Worst project (by exposure)",
     f'=IFERROR(INDEX($J${HLP_FIRST}:$J${HLP_LAST},{WORST}),"")', None),
    ("  its milestone hit rate",
     f'=IFERROR(INDEX(Dashboard!$J${PROW_FIRST}:$J${PROW_LAST},{WORST}),"")', PCT),
    ("  its budget used",
     f'=IFERROR(INDEX(Dashboard!$F${PROW_FIRST}:$F${PROW_LAST},{WORST}),"")', PCT),
    ("  its high risks",
     f'=IFERROR(INDEX(Dashboard!$P${PROW_FIRST}:$P${PROW_LAST},{WORST}),"")', INT),
    ("  its calculated RAG",
     f'=IFERROR(INDEX(Dashboard!$S${PROW_FIRST}:$S${PROW_LAST},{WORST}),"")', None),
    ("Prior month", '=DATE(YEAR($C$5),MONTH($C$5)-1,1)', MONTH),
    ("Prior ms planned",
     f'=SUMIFS({ML}$C${LOG_FIRST}:$C${LOG_LAST},{ML}$A${LOG_FIRST}:$A${LOG_LAST},$K$54)', INT),
    ("Prior ms achieved",
     f'=SUMIFS({ML}$D${LOG_FIRST}:$D${LOG_LAST},{ML}$A${LOG_FIRST}:$A${LOG_LAST},$K$54)', INT),
    ("Prior hit rate", '=IF($K$55=0,"",$K$56/$K$55)', PCT),
    ("Prior spend",
     f'=SUMIFS({ML}$G${LOG_FIRST}:$G${LOG_LAST},{ML}$A${LOG_FIRST}:$A${LOG_LAST},$K$54)', CUR),
    ("Milestones behind",
     f'=SUM(Dashboard!$H${PROW_FIRST}:$H${PROW_LAST})'
     f'-SUM(Dashboard!$I${PROW_FIRST}:$I${PROW_LAST})', INT),
    ("Deliverables not accepted",
     f'=SUM(Dashboard!$K${PROW_FIRST}:$K${PROW_LAST})'
     f'-SUM(Dashboard!$L${PROW_FIRST}:$L${PROW_LAST})', INT),
    ("Approved CRs, total",
     f'=SUM(Dashboard!$N${PROW_FIRST}:$N${PROW_LAST})', INT),
    ("Budget headroom", '=Dashboard!$I$8-Dashboard!$K$8', CUR),
    ("Milestones planned, total",
     f'=SUM(Dashboard!$H${PROW_FIRST}:$H${PROW_LAST})', INT),
    ("Milestones achieved, total",
     f'=SUM(Dashboard!$I${PROW_FIRST}:$I${PROW_LAST})', INT),
    ("Deliverables due, total",
     f'=SUM(Dashboard!$K${PROW_FIRST}:$K${PROW_LAST})', INT),
    ("Deliverables accepted, total",
     f'=SUM(Dashboard!$L${PROW_FIRST}:$L${PROW_LAST})', INT),
]
for i, (lbl, formula, fmt) in enumerate(SCALARS):
    rr = 46 + i
    a = brief.cell(row=rr, column=10, value=lbl)
    a.font = Font(name=FONT, size=8, color=SLATE)
    a.border = BOX
    body(brief.cell(row=rr, column=11, value=formula), fmt=fmt, align="center")

# ================================================================== LISTS =====
ls = wb.create_sheet("Lists")
ls.sheet_properties.tabColor = GREY
title_block(ls, "  Dropdown lists",
            "  Extend a column and the matching dropdown picks it up. Keep RAG wording as "
            "Green / Amber / Red and cadence as Daily / Monthly - the formulas match on them.", 6)
header_row(ls, 3, ["RAG", "Deliverable Status", "Risk/Issue Status", "Type",
                   "Score 1-5", "Cadence"])
widths(ls, {"A": 14, "B": 20, "C": 20, "D": 12, "E": 12, "F": 14})
for col, vals in enumerate([LIST_RAG, LIST_DSTATUS, LIST_RSTATUS, LIST_RTYPE,
                            LIST_SCORE, LIST_CADENCE], start=1):
    for i, v in enumerate(vals):
        body(ls.cell(row=4 + i, column=col, value=v), color=BLUE_INPUT, align="center")

DVS = [
    (day, f"M{DAY_FIRST}:M{DAY_LAST}", f"=Lists!$A$4:$A${3 + len(LIST_RAG)}"),
    (log, f"M{LOG_FIRST}:M{LOG_LAST}", f"=Lists!$A$4:$A${3 + len(LIST_RAG)}"),
    (dl, f"J{DEL_FIRST}:J{DEL_LAST}", f"=Lists!$B$4:$B${3 + len(LIST_DSTATUS)}"),
    (rk, f"K{RSK_FIRST}:K{RSK_LAST}", f"=Lists!$C$4:$C${3 + len(LIST_RSTATUS)}"),
    (rk, f"C{RSK_FIRST}:C{RSK_LAST}", f"=Lists!$D$4:$D${3 + len(LIST_RTYPE)}"),
    (rk, f"E{RSK_FIRST}:F{RSK_LAST}", f"=Lists!$E$4:$E${3 + len(LIST_SCORE)}"),
    (db, "C3", f"=Lists!$F$4:$F${3 + len(LIST_CADENCE)}"),
]
for sheet, rng, src in DVS:
    dv = DataValidation(type="list", formula1=src, allow_blank=True, showErrorMessage=True)
    dv.error = "Pick a value from the list on the Lists tab."
    dv.errorTitle = "Not a valid entry"
    sheet.add_data_validation(dv)
    dv.add(rng)

# project-ID dropdowns driven by the TOR Register
pid_src = f"='TOR Register'!$A${TOR_FIRST}:$A${TOR_LAST}"
for sheet, rng in ((day, f"B{DAY_FIRST}:B{DAY_LAST}"),
                   (log, f"B{LOG_FIRST}:B{LOG_LAST}"),
                   (dl, f"B{DEL_FIRST}:B{DEL_LAST}"),
                   (rk, f"B{RSK_FIRST}:B{RSK_LAST}")):
    dv = DataValidation(type="list", formula1=pid_src, allow_blank=True, showErrorMessage=False)
    sheet.add_data_validation(dv)
    dv.add(rng)

# --- finalisation: Arial everywhere, including merged continuation and blank cells
try:
    wb._named_styles["Normal"].font = Font(name=FONT, size=10)
except Exception:
    pass

for sheet in wb.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            f = cell.font
            if f is not None and f.name != FONT:
                cell.font = Font(name=FONT, size=f.size or 10, bold=f.bold,
                                 italic=f.italic, color=f.color)

# CEO Brief sits second, right after the Read Me, and is what the file opens on.
_sheets = wb._sheets
_sheets.insert(1, _sheets.pop(_sheets.index(brief)))

wb.active = wb.sheetnames.index("CEO Brief")
wb.save(OUT)
print("wrote", OUT)
print(f"  daily example rows   : {len(daily_rows)} over {len(DAILY_DAYS)} working days "
      f"({DAILY_DAYS[0]} to {DAILY_DAYS[-1]})")
print(f"  monthly example rows : {len(monthly_rows)} over {len(MONTHS)} months")
